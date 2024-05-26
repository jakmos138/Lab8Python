import argparse
import sys
import os
import re
import csv
from datetime import datetime
import openpyxl
from openpyxl.styles import Font

# URL of the dataset
# URL: https://www.kaggle.com/datasets/vishnurajyadav12/movies-dataset-tmdb


class movie:

    def __init__(self, id, title, overview,
                 releaseDate, popularity,
                 voteAverage, voteCount):
        self.id = id
        self.title = title
        self.overview = overview
        self.releaseDate = releaseDate
        self.popularity = popularity
        self.voteAverage = voteAverage
        self.voteCount = voteCount

    def __str__(self):
        return f"""Movie Id: {self.id}
                Title: {self.title}
                Overview: {self.overview}
                Release Date: {self.releaseDate}
                Popularity: {self.popularity}
                Vote Average: {self.voteAverage}
                Vote Count: {self.voteCount}"""


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Python app to analyze CSV dataset and generate report"
    )
    parser.add_argument("dataset", help="Path to the dataset CSV file")
    parser.add_argument(
        "-o",
        "--output",
        help="Name of the Excel file to be generated"
    )
    return parser.parse_args()


def validate_file(filename):
    if not filename.endswith('.csv'):
        sys.exit("Error: The file must have a .csv extension.")
    if not os.path.isfile(filename):
        sys.exit(f"Error: The file {filename} does not exist.")


def validate_output_file(filename):
    if filename and not filename.endswith('.xlsx'):
        sys.exit("Error: The output file must have a .xlsx extension.")


def parse_csv_file(filename):
    movies = []
    with open(filename, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(
            csvfile,
            delimiter=',',
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL
        )
        next(reader)  # Skip the header row
        for row in reader:
            try:
                id = int(row[1])
                title = row[2]
                overview = row[3]
                release_date = datetime.strptime(row[4], '%Y-%m-%d')
                popularity = float(row[5])
                vote_average = float(row[6])
                vote_count = int(row[7])
                entry = movie(id, title, overview,
                              release_date, popularity,
                              vote_average, vote_count)
                movies.append(entry)
            except (IndexError, ValueError) as e:
                print("Error parsing row:", row)
                print("Exception:", e)
    return movies


def average_vote_count(movies):
    # Calculate the average vote count for all movies in the dataset.
    total_vote_count = sum(movie.voteCount for movie in movies)
    average = total_vote_count / len(movies)
    return average


def group_by_release_year(movies):
    # Calculate the total number of movies released each year.
    movies_by_year = {}
    for movie in movies:
        release_year = movie.releaseDate.year
        if release_year not in movies_by_year:
            movies_by_year[release_year] = 1
        else:
            movies_by_year[release_year] += 1
    return movies_by_year


def total_movies_and_popularity(movies):
    # Calculate the the total popularity score of all movies combined.
    total_movies = len(movies)
    total_popularity = sum(movie.popularity for movie in movies)
    return total_movies, total_popularity


def run():
    args = parse_arguments()
    validate_file(args.dataset)
    validate_output_file(args.output)
    movies = parse_csv_file(args.dataset)

    # Perform operations
    avg_vote_count = average_vote_count(movies)
    movies_by_year = group_by_release_year(movies)
    total_movies, total_popularity = total_movies_and_popularity(movies)

    # Save results to Excel file if output file is provided
    if args.output:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movie Report"

        # Write results to Excel file
        ws['A1'] = "Average Vote Count:"
        ws['A2'] = avg_vote_count

        ws['D1'] = "Movies by Release Year:"
        for i, (year, count) in enumerate(movies_by_year.items(), start=2):
            ws[f'D{i}'] = year
            ws[f'E{i}'] = count

        ws['A4'] = "Total Movies:"
        ws['A5'] = total_movies

        ws['A7'] = "Total Popularity Score:"
        ws['A8'] = total_popularity

        # Apply text formatting
        header_cells = ['A1', 'D1', 'A4', 'A7']
        for cell in header_cells:
            ws[cell].font = Font(name='Arial', size=12,
                                 bold=True, color="ff4444aa")

        # Save Excel file
        wb.save(args.output)
        print(f"Results saved to {args.output}")

    # Display summary results if output file is not provided
    else:
        print("Average Vote Count:", avg_vote_count)
        print("Movies by Release Year:", movies_by_year)
        print("Total Movies:", total_movies)
        print("Total Popularity Score:", total_popularity)


if __name__ == "__main__":
    run()
